from __future__ import annotations

import base64
import io
import json
import logging
import re
from collections import OrderedDict
from dataclasses import dataclass
from datetime import datetime, timedelta, timezone
from http import HTTPStatus
from typing import Any, Dict, Iterable, List, Optional, Tuple

from database.connector_bot import (
    add_to_blacklist,
    fetch_audit_log_entries,
    fetch_code_statistics,
    fetch_code_statistics_insights,
    fetch_code_statistics_for_export,
    fetch_working_hours_entries,
    get_blacklist,
    get_blacklist_with_meta,
    get_connection,
    get_setting,
    get_inventory_name_map,
    record_audit_event,
    remove_from_blacklist,
    save_working_hours_entries,
    set_setting,
)
from handlers.inventory import refresh_inventory_cache_once
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from . import runtime

LOGGER = logging.getLogger(__name__)

try:
    from privateTelegram.config.settings import (  # type: ignore
        load_settings as _private_reload_settings,
        save_settings as _private_save_settings,
        settings as _private_settings,
    )
except Exception:  # pragma: no cover - optional dependency
    _private_reload_settings = None
    _private_save_settings = None
    _private_settings = None

try:  # pragma: no cover - optional dependency
    from privateTelegram.metrics.tracker import get_snapshot as _private_metrics_snapshot  # type: ignore
except Exception:
    _private_metrics_snapshot = None

try:  # pragma: no cover - optional dependency
    from privateTelegram.cache import refresh_cache_once as _private_refresh_cache  # type: ignore
except Exception:
    _private_refresh_cache = None

UTC = timezone.utc

PERSIAN_MONTHS = [
    "ژانویه",
    "فوریه",
    "مارس",
    "آوریل",
    "مه",
    "ژوئن",
    "ژوئیه",
    "اوت",
    "سپتامبر",
    "اکتبر",
    "نوامبر",
    "دسامبر",
]

DEFAULT_COMMANDS = [
    {
        "id": "cmd-start",
        "command": "/start",
        "description": "آغاز مکالمه با کاربر و نمایش منو.",
        "enabled": True,
        "lastUsedISO": None,
    },
    {
        "id": "cmd-help",
        "command": "/help",
        "description": "نمایش راهنمای استفاده از ربات.",
        "enabled": True,
        "lastUsedISO": None,
    },
]

COMMANDS_KEY = "panel_commands_v1"
PLATFORMS_KEY = "panel_platforms_v1"
CACHE_KEY = "inventory_cache_last_refresh"
TIMEZONE_KEY = "working_timezone"
LUNCH_START_KEY = "lunch_start"
LUNCH_END_KEY = "lunch_end"
QUERY_LIMIT_KEY = "query_limit"
DELIVERY_BEFORE_KEY = "delivery_before"
DELIVERY_AFTER_KEY = "delivery_after"
CHANGEOVER_KEY = "changeover_hour"
AUDIT_LOG_KEY = "panel_audit_log_v1"
BLOCKLIST_META_KEY = "panel_blocklist_meta_v1"

WORKING_DAY_ORDER = [5, 6, 0, 1, 2, 3, 4]  # Saturday → Friday (Python weekday numbering)

MAX_AUDIT_LOG_ENTRIES = 200

class ControlPanelError(Exception):
    """Raised for validation or domain errors that should be surfaced to the API."""

    def __init__(self, message: str, status: int = 400) -> None:
        super().__init__(message)
        self.message = message
        self.status = status


@dataclass
class Metrics:
    totals: Dict[str, int]
    monthly: List[Dict[str, Any]]
    cache: Dict[str, Optional[str]]
    status: Dict[str, Any]


def _now_iso() -> str:
    return datetime.now(UTC).isoformat()


def _normalize_time(value: Optional[str], field_name: str) -> str:
    if value is None:
        return ""
    text = str(value).strip()
    if not text:
        return ""
    try:
        datetime.strptime(text, "%H:%M")
    except Exception:
        raise ControlPanelError(f"زمان {field_name} باید در قالب HH:MM باشد.")
    return text


def _safe_int(value: Optional[str]) -> Optional[int]:
    if value is None:
        return None
    try:
        return int(str(value).strip())
    except Exception:
        return None


def _parse_date_field(value: Any, field_name: str, *, end_of_day: bool = False) -> Optional[datetime]:
    if value in (None, ""):
        return None
    text = str(value).strip()
    if not text:
        return None
    try:
        if len(text) == 10:
            parsed = datetime.strptime(text, "%Y-%m-%d")
        else:
            parsed = datetime.fromisoformat(text)
    except Exception:
        raise ControlPanelError(f"تاریخ {field_name} معتبر نیست. از قالب YYYY-MM-DD استفاده کنید.")
    if parsed.tzinfo is not None:
        parsed = parsed.astimezone(UTC).replace(tzinfo=None)
    if end_of_day:
        return parsed.replace(hour=23, minute=59, second=59, microsecond=999000)
    return parsed.replace(hour=0, minute=0, second=0, microsecond=0)


def _sanitize_excel_filename(value: Any) -> str:
    base = str(value or "").strip()
    if not base:
        base = "گزارش-آمار"
    cleaned = re.sub(r"[\\/:*?\"<>|]+", "-", base)
    cleaned = re.sub(r"\.(xlsx?|XLSX?)$", "", cleaned, flags=re.IGNORECASE)
    safe = cleaned or "گزارش-آمار"
    return f"{safe}.xlsx"


def _load_json_setting(key: str, default: Any) -> Any:
    raw = get_setting(key)
    if not raw:
        return default
    try:
        return json.loads(raw)
    except Exception:
        LOGGER.warning("Failed to decode JSON setting %s", key)
        return default


def _collect_private_telegram_metrics() -> Tuple[int, Dict[Tuple[int, int], int]]:
    if not _private_metrics_snapshot:
        return 0, {}

    try:
        total, entries = _private_metrics_snapshot(limit=24)
    except Exception:
        LOGGER.debug("Failed to read private Telegram metrics snapshot", exc_info=True)
        return 0, {}

    buckets: Dict[Tuple[int, int], int] = {}
    for year, month, count in entries:
        try:
            key = (int(year), int(month))
            buckets[key] = buckets.get(key, 0) + int(count)
        except Exception:
            continue

    return int(total), buckets


def _save_json_setting(key: str, value: Any) -> None:
    try:
        set_setting(key, json.dumps(value, ensure_ascii=False))
    except Exception:
        LOGGER.exception("Failed to persist JSON setting %s", key)
        raise ControlPanelError("ذخیره‌سازی تنظیمات امکان‌پذیر نبود. دوباره تلاش کنید.", status=500)


def _load_blocklist_meta() -> Dict[str, str]:
    data = _load_json_setting(BLOCKLIST_META_KEY, {})
    if isinstance(data, dict):
        result: Dict[str, str] = {}
        for key, value in data.items():
            key_text = str(key).strip()
            value_text = str(value).strip()
            if key_text and value_text:
                result[key_text] = value_text
        return result
    return {}


def _save_blocklist_meta(meta: Dict[str, str]) -> None:
    try:
        set_setting(BLOCKLIST_META_KEY, json.dumps(meta, ensure_ascii=False))
    except Exception:
        LOGGER.warning("Failed to persist blocklist metadata", exc_info=True)


def _remember_blocklist_timestamp(user_id: int, timestamp: str) -> None:
    key = str(user_id)
    if not timestamp:
        return
    meta = _load_blocklist_meta()
    if meta.get(key) == timestamp:
        return
    meta[key] = timestamp
    _save_blocklist_meta(meta)


def _forget_blocklist_timestamp(user_id: int) -> None:
    key = str(user_id)
    meta = _load_blocklist_meta()
    if key not in meta:
        return
    meta.pop(key, None)
    _save_blocklist_meta(meta)


def _load_audit_log_fallback() -> List[Dict[str, Any]]:
    data = _load_json_setting(AUDIT_LOG_KEY, [])
    if not isinstance(data, list):
        return []

    entries: List[Dict[str, Any]] = []
    for item in data:
        if not isinstance(item, dict):
            continue
        message = str(item.get("message") or "").strip()
        timestamp = item.get("timestamp") or ""
        if not message or not timestamp:
            continue
        entry = {
            "id": str(item.get("id") or f"log-{len(entries) + 1}"),
            "timestamp": timestamp,
            "message": message,
        }
        if item.get("actor"):
            entry["actor"] = str(item.get("actor"))
        if item.get("details"):
            entry["details"] = item.get("details")
        entries.append(entry)
        if len(entries) >= MAX_AUDIT_LOG_ENTRIES:
            break
    return entries


def _persist_audit_log_fallback(entries: List[Dict[str, Any]]) -> None:
    try:
        _save_json_setting(AUDIT_LOG_KEY, entries[:MAX_AUDIT_LOG_ENTRIES])
    except ControlPanelError:
        # Audit log persistence failures should not block the primary action.
        LOGGER.debug("Failed to persist audit trail entry", exc_info=True)


def _append_audit_event(message: str, *, actor: str = "کنترل‌پنل", details: Optional[Any] = None) -> None:
    entry = {
        "id": f"log-{int(datetime.now().timestamp() * 1000)}",
        "timestamp": _now_iso(),
        "message": message,
        "actor": actor,
    }
    if details:
        entry["details"] = details
    try:
        record_audit_event(message, actor=actor, details=details)
        return
    except Exception:
        LOGGER.debug("Falling back to settings-based audit trail persistence", exc_info=True)

    entries = _load_audit_log_fallback()
    entries.insert(0, entry)
    _persist_audit_log_fallback(entries)


def _ensure_private_settings() -> Dict[str, Any]:
    if (
        _private_settings is None
        or _private_save_settings is None
        or _private_reload_settings is None
    ):
        raise ControlPanelError(
            "ماژول تلگرام خصوصی در دسترس نیست.", status=HTTPStatus.SERVICE_UNAVAILABLE
        )
    try:
        _private_reload_settings()
    except Exception as exc:  # pragma: no cover - defensive I/O guard
        LOGGER.warning("Failed to reload private Telegram settings: %s", exc)
        raise ControlPanelError("خواندن تنظیمات تلگرام خصوصی ممکن نشد.", status=500)
    return _private_settings


def _snapshot_private_settings() -> Dict[str, Any]:
    data = _ensure_private_settings()
    try:
        return json.loads(json.dumps(data))
    except Exception:
        # Fallback to shallow copy if encoding fails for unexpected types
        return dict(data)


def _default_private_platform_enabled() -> bool:
    if _private_settings is None:
        return True

    try:
        data = _ensure_private_settings()
    except ControlPanelError:
        return True
    except Exception:
        LOGGER.debug(
            "Failed to determine private Telegram default platform state", exc_info=True
        )
        return True

    return bool(data.get("enabled", True))


def _normalize_group_list(value: Any, field_name: str) -> List[int]:
    if value is None or value == "":
        return []
    items: List[int] = []
    source: Iterable[Any]
    if isinstance(value, str):
        source = [token for token in re.split(r"[\s,\n]+", value) if token.strip()]
    elif isinstance(value, Iterable):
        source = value
    else:
        raise ControlPanelError(f"ساختار {field_name} نامعتبر است.")

    for item in source:
        if item in (None, ""):
            continue
        try:
            items.append(int(str(item).strip()))
        except Exception:
            raise ControlPanelError(f"مقادیر {field_name} باید عددی باشند.")
    return items


def _normalize_time_range(payload: Any, start_label: str, end_label: str) -> Dict[str, str]:
    data = payload or {}
    if not isinstance(data, dict):
        raise ControlPanelError("ساختار بازه زمانی نامعتبر است.")
    start = _normalize_time(data.get("start"), start_label)
    end = _normalize_time(data.get("end"), end_label)
    return {"start": start, "end": end}


def _format_month_label(year: int, month: int) -> str:
    index = max(1, min(month, 12)) - 1
    month_name = PERSIAN_MONTHS[index]
    return f"{month_name} {year}"


def _is_globally_enabled() -> bool:
    return (get_setting("enabled") or "true").strip().lower() == "true"


def _table_exists(cur, table_name: str) -> bool:
    try:
        schema = None
        name = table_name
        if "." in table_name:
            schema, name = table_name.split(".", 1)
        candidates = {name.strip("[]"), name, name.upper(), name.lower()}
        for candidate in candidates:
            if not candidate:
                continue
            params = [candidate]
            schema_clause = ""
            if schema:
                schema_clause = " AND TABLE_SCHEMA = ?"
                params.append(schema)
            row = cur.execute(
                f"SELECT 1 FROM INFORMATION_SCHEMA.TABLES WHERE TABLE_NAME = ?{schema_clause}",
                params,
            ).fetchone()
            if row:
                return True
    except Exception:
        return False
    return False


def _column_exists(cur, table_name: str, column_name: str) -> bool:
    try:
        schema = None
        table = table_name
        if "." in table_name:
            schema, table = table_name.split(".", 1)
        params = [table.strip("[]"), column_name]
        schema_clause = ""
        if schema:
            schema_clause = " AND TABLE_SCHEMA = ?"
            params.insert(1, schema)
        row = cur.execute(
            f"SELECT 1 FROM INFORMATION_SCHEMA.COLUMNS WHERE TABLE_NAME = ?{schema_clause} AND COLUMN_NAME = ?",
            params,
        ).fetchone()
        return bool(row)
    except Exception:
        return False


def _collect_metrics(cur) -> Tuple[Dict[str, int], List[Dict[str, Any]]]:
    totals = {"telegram": 0, "whatsapp": 0, "privateTelegram": 0, "all": 0}
    monthly_map: OrderedDict = OrderedDict()

    def record(
        year: int,
        month: int,
        telegram: int = 0,
        whatsapp: int = 0,
        private_telegram: int = 0,
    ) -> None:
        key = (year, month)
        bucket = monthly_map.setdefault(
            key, {"telegram": 0, "whatsapp": 0, "privateTelegram": 0}
        )
        bucket["telegram"] += telegram
        bucket["whatsapp"] += whatsapp
        bucket["privateTelegram"] += private_telegram

    if _column_exists(cur, "message_log", "platform"):
        rows = cur.execute(
            """
            SELECT
                YEAR(timestamp) AS y,
                MONTH(timestamp) AS m,
                SUM(CASE WHEN LOWER(platform) = 'telegram' THEN 1 ELSE 0 END) AS telegram_cnt,
                SUM(CASE WHEN LOWER(platform) = 'whatsapp' THEN 1 ELSE 0 END) AS whatsapp_cnt
            FROM message_log
            WHERE timestamp >= DATEADD(MONTH, -11, GETDATE())
            GROUP BY YEAR(timestamp), MONTH(timestamp)
            ORDER BY y, m
            """
        ).fetchall()
        for row in rows:
            year, month = int(row[0]), int(row[1])
            telegram_cnt = int(row[2] or 0)
            whatsapp_cnt = int(row[3] or 0)
            totals["telegram"] += telegram_cnt
            totals["whatsapp"] += whatsapp_cnt
            record(year, month, telegram_cnt, whatsapp_cnt)
    else:
        rows = cur.execute(
            """
            SELECT
                YEAR(timestamp) AS y,
                MONTH(timestamp) AS m,
                COUNT(*) AS cnt
            FROM message_log
            WHERE timestamp >= DATEADD(MONTH, -11, GETDATE())
            GROUP BY YEAR(timestamp), MONTH(timestamp)
            ORDER BY y, m
            """
        ).fetchall()
        for row in rows:
            year, month = int(row[0]), int(row[1])
            count = int(row[2] or 0)
            totals["telegram"] += count
            record(year, month, telegram=count)

    whatsapp_sources = [
        "wa_message_log",
        "whatsapp_message_log",
        "whatsapp_log",
        "message_log_whatsapp",
    ]
    for table in whatsapp_sources:
        if _table_exists(cur, table):
            rows = cur.execute(
                f"""
                SELECT
                    YEAR(timestamp) AS y,
                    MONTH(timestamp) AS m,
                    COUNT(*) AS cnt
                FROM {table}
                WHERE timestamp >= DATEADD(MONTH, -11, GETDATE())
                GROUP BY YEAR(timestamp), MONTH(timestamp)
                ORDER BY y, m
                """
            ).fetchall()
            for row in rows:
                year, month = int(row[0]), int(row[1])
                count = int(row[2] or 0)
                totals["whatsapp"] += count
                record(year, month, whatsapp=count)
            break

    private_total, private_monthly = _collect_private_telegram_metrics()
    totals["privateTelegram"] += private_total
    for (year, month), count in private_monthly.items():
        record(year, month, private_telegram=count)

    totals["all"] = (
        totals["telegram"] + totals["whatsapp"] + totals["privateTelegram"]
    )

    monthly: List[Dict[str, Any]] = []
    for (year, month), values in monthly_map.items():
        telegram = values.get("telegram", 0)
        whatsapp = values.get("whatsapp", 0)
        private_telegram = values.get("privateTelegram", 0)
        monthly.append(
            {
                "month": _format_month_label(year, month),
                "telegram": telegram,
                "whatsapp": whatsapp,
                "privateTelegram": private_telegram,
                "all": telegram + whatsapp + private_telegram,
            }
        )

    return totals, monthly


def _aggregate_metrics_from_db() -> Metrics:
    status = _build_status_snapshot()

    fallback = False
    try:
        with get_connection() as conn:
            cur = conn.cursor()
            totals, monthly = _collect_metrics(cur)
    except Exception as exc:
        fallback = True
        LOGGER.warning("Failed to aggregate metrics from database: %s", exc)
        monthly = _build_mock_monthly()
        totals = _build_mock_totals()

    cache_info = {
        "lastUpdatedISO": get_setting(CACHE_KEY) or _now_iso(),
        "usingFallback": fallback,
    }
    status["dataSource"] = "fallback" if fallback else "live"
    status["usingFallback"] = fallback
    return Metrics(totals=totals, monthly=monthly, cache=cache_info, status=status)


def _build_mock_totals() -> Dict[str, int]:
    telegram = 320
    whatsapp = 185
    private_telegram = 210
    return {
        "telegram": telegram,
        "whatsapp": whatsapp,
        "privateTelegram": private_telegram,
        "all": telegram + whatsapp + private_telegram,
    }


def _build_mock_monthly() -> List[Dict[str, Any]]:
    today = datetime.now()
    results: List[Dict[str, Any]] = []
    for offset in range(11, -1, -1):
        current = today - timedelta(days=30 * offset)
        year = current.year
        month = current.month
        label = _format_month_label(year, month)
        telegram = 50 + (offset * 3)
        whatsapp = 35 + (offset * 2)
        private_telegram = 40 + (offset * 2)
        results.append(
            {
                "month": label,
                "telegram": telegram,
                "whatsapp": whatsapp,
                "privateTelegram": private_telegram,
                "all": telegram + whatsapp + private_telegram,
            }
        )
    return results


def _load_platform_settings(enabled: bool) -> Dict[str, bool]:
    stored = _load_json_setting(PLATFORMS_KEY, None)
    defaults = {
        "telegram": enabled,
        "whatsapp": True,
        "privateTelegram": _default_private_platform_enabled(),
    }
    merged = {**defaults}
    if isinstance(stored, dict):
        for key, value in stored.items():
            if key in merged:
                merged[key] = bool(value)
    if not enabled:
        return {name: False for name in merged}
    return merged


def _merge_platform_flags(current: Dict[str, bool], overrides: Dict[str, Any]) -> Dict[str, bool]:
    merged = {**current}
    for key, value in overrides.items():
        if key in merged:
            merged[key] = bool(value)
    return merged


def _order_weekly_items(items: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    order_map = {day: index for index, day in enumerate(WORKING_DAY_ORDER)}
    return sorted(
        items,
        key=lambda item: order_map.get(int(item.get("day", -1)), len(WORKING_DAY_ORDER)),
    )


def _legacy_weekly_schedule() -> List[Dict[str, Any]]:
    general_open = (get_setting("working_start") or "08:00").strip()
    general_close = (get_setting("working_end") or "18:00").strip()
    th_open = (get_setting("thursday_start") or general_open).strip()
    th_close = (get_setting("thursday_end") or general_close).strip()
    friday_disabled = (get_setting("disable_friday") or "true").lower() == "true"

    legacy: List[Dict[str, Any]] = []
    for day in WORKING_DAY_ORDER:
        if day == 3:  # Thursday
            open_time = th_open or None
            close_time = th_close or None
        elif day == 4:  # Friday
            if friday_disabled:
                open_time = None
                close_time = None
            else:
                open_time = general_open or None
                close_time = general_close or None
        else:
            open_time = general_open or None
            close_time = general_close or None
        legacy.append({"day": day, "open": open_time, "close": close_time})
    return legacy


def _build_weekly_schedule() -> List[Dict[str, Any]]:
    try:
        entries = fetch_working_hours_entries()
    except Exception:
        LOGGER.debug("Falling back to legacy working hours settings", exc_info=True)
        return _order_weekly_items(_legacy_weekly_schedule())

    weekly: List[Dict[str, Any]] = []
    seen: set[int] = set()
    for entry in entries:
        try:
            day = int(entry.get("day"))
        except Exception:
            continue
        if day < 0 or day > 6 or day in seen:
            continue
        open_value = entry.get("open")
        close_value = entry.get("close")
        if entry.get("closed") or not (open_value and close_value):
            open_value = None
            close_value = None
        weekly.append({"day": day, "open": open_value, "close": close_value})
        seen.add(day)

    if len(weekly) < len(WORKING_DAY_ORDER):
        fallback_map = {item["day"]: item for item in _legacy_weekly_schedule()}
        for day in WORKING_DAY_ORDER:
            if day not in seen and day in fallback_map:
                weekly.append(fallback_map[day])

    return _order_weekly_items(weekly)


def _normalize_weekly_payload(payload: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    normalized: Dict[int, Dict[str, Any]] = {}
    for item in payload:
        if not isinstance(item, dict):
            continue
        try:
            day = int(item.get("day"))
        except Exception:
            raise ControlPanelError("شناسه روز نامعتبر است.")
        if day < 0 or day > 6:
            raise ControlPanelError("شناسه روز باید بین 0 و 6 باشد.")

        open_raw = item.get("open")
        close_raw = item.get("close")
        if open_raw in (None, "") or close_raw in (None, ""):
            normalized[day] = {"day": day, "open": None, "close": None, "closed": True}
            continue

        open_text = _normalize_time(open_raw, "شروع کار")
        close_text = _normalize_time(close_raw, "پایان کار")
        try:
            open_time = datetime.strptime(open_text, "%H:%M").time()
            close_time = datetime.strptime(close_text, "%H:%M").time()
        except Exception:
            raise ControlPanelError("ساعت وارد شده نامعتبر است.")
        if open_time >= close_time:
            raise ControlPanelError("ساعت پایان باید بعد از ساعت شروع باشد.")

        normalized[day] = {
            "day": day,
            "open": open_text,
            "close": close_text,
            "closed": False,
        }

    if not normalized:
        raise ControlPanelError("هیچ ساعتی برای ذخیره ارسال نشده است.")

    return _order_weekly_items(list(normalized.values()))


def _sync_legacy_working_settings(entries: Iterable[Dict[str, Any]]) -> None:
    entries_map = {int(item["day"]): item for item in entries if "day" in item}
    general_candidates = [0, 1, 2, 5, 6, 3]
    general = next(
        (
            entries_map[day]
            for day in general_candidates
            if day in entries_map
            and entries_map[day].get("open")
            and entries_map[day].get("close")
        ),
        None,
    )
    if general:
        set_setting("working_start", general.get("open") or "09:00")
        set_setting("working_end", general.get("close") or "18:00")

    thursday = entries_map.get(3)
    if thursday and thursday.get("open") and thursday.get("close"):
        set_setting("thursday_start", thursday.get("open") or "")
        set_setting("thursday_end", thursday.get("close") or "")
    else:
        set_setting("thursday_start", "")
        set_setting("thursday_end", "")

    friday = entries_map.get(4)
    if not friday or friday.get("closed") or not (friday.get("open") and friday.get("close")):
        set_setting("disable_friday", "true")
    else:
        set_setting("disable_friday", "false")


def _build_status_snapshot() -> Dict[str, Any]:
    enabled = _is_globally_enabled()
    weekly = _build_weekly_schedule()
    platforms = _load_platform_settings(enabled)
    message = "ربات فعال و آماده پاسخ‌گویی است." if enabled else "ربات غیرفعال است."

    timezone_value = get_setting(TIMEZONE_KEY) or "Asia/Tehran"

    lunch_start = get_setting(LUNCH_START_KEY) or ""
    lunch_end = get_setting(LUNCH_END_KEY) or ""
    query_limit = _safe_int(get_setting(QUERY_LIMIT_KEY))
    delivery_before = get_setting(DELIVERY_BEFORE_KEY) or ""
    delivery_after = get_setting(DELIVERY_AFTER_KEY) or ""
    changeover_hour = get_setting(CHANGEOVER_KEY) or ""

    operations = {
        "lunchBreak": {
            "start": lunch_start or None,
            "end": lunch_end or None,
        },
        "queryLimit": query_limit,
        "delivery": {
            "before": delivery_before,
            "after": delivery_after,
            "changeover": changeover_hour or None,
        },
    }

    return {
        "active": enabled,
        "message": message,
        "workingHours": {
            "timezone": timezone_value,
            "weekly": weekly,
        },
        "platforms": platforms,
        "operations": operations,
        "dataSource": "live",
    }


def get_metrics() -> Dict[str, Any]:
    metrics = _aggregate_metrics_from_db()
    return {
        "totals": metrics.totals,
        "monthly": metrics.monthly,
        "cache": metrics.cache,
        "status": metrics.status,
    }


def _build_code_statistics_insights(range_key: str, search_value: str) -> Optional[Dict[str, Any]]:
    search_arg = (search_value or "").strip() or None
    try:
        raw = fetch_code_statistics_insights(range_key=range_key, search=search_arg)
    except Exception as exc:
        LOGGER.debug("Failed to compute code statistics insights: %s", exc, exc_info=True)
        return None

    total_requests = int(raw.get("totalRequests", 0) or 0)
    unique_codes = int(raw.get("uniqueCodes", 0) or 0)
    average_per_day = raw.get("averagePerDay")
    try:
        average_per_day_value = round(float(average_per_day), 2)
    except Exception:
        average_per_day_value = float(total_requests)

    top_codes: List[Dict[str, Any]] = []
    for item in raw.get("topCodes") or []:
        try:
            share_value = round(float(item.get("sharePercent", 0) or 0), 2)
        except Exception:
            share_value = 0.0
        top_codes.append(
            {
                "code": str(item.get("code") or "").strip(),
                "partName": str(item.get("partName") or "-").strip() or "-",
                "requestCount": int(item.get("requestCount", 0) or 0),
                "sharePercent": share_value,
            }
        )

    daily_trend: List[Dict[str, Any]] = []
    for item in raw.get("dailyTrend") or []:
        date_value = str(item.get("date") or "").strip()
        if not date_value:
            continue
        daily_trend.append(
            {
                "date": date_value,
                "requestCount": int(item.get("requestCount", 0) or 0),
            }
        )

    platform_breakdown: List[Dict[str, Any]] = []
    for item in raw.get("platformBreakdown") or []:
        try:
            share_value = round(float(item.get("sharePercent", 0) or 0), 2)
        except Exception:
            share_value = 0.0
        platform_breakdown.append(
            {
                "platform": str(item.get("platform") or "unknown").strip() or "unknown",
                "requestCount": int(item.get("requestCount", 0) or 0),
                "sharePercent": share_value,
            }
        )

    return {
        "totalRequests": total_requests,
        "uniqueCodes": unique_codes,
        "firstRequestISO": raw.get("firstRequestISO"),
        "lastRequestISO": raw.get("lastRequestISO"),
        "activeDays": int(raw.get("activeDays", len(daily_trend)) or len(daily_trend)),
        "averagePerDay": average_per_day_value,
        "topCodes": top_codes,
        "dailyTrend": daily_trend,
        "platformBreakdown": platform_breakdown,
    }


def get_code_statistics(
    *,
    range_key: str,
    sort_order: str,
    page: int,
    page_size: int,
    search: Optional[str] = None,
) -> Dict[str, Any]:
    search_value = (search or "").strip()
    try:
        items, total = fetch_code_statistics(
            range_key=range_key,
            sort_order=sort_order,
            page=page,
            page_size=page_size,
            search=search_value or None,
        )
    except Exception as exc:
        LOGGER.warning("Failed to fetch code statistics: %s", exc)
        raise ControlPanelError("دریافت آمار کدها امکان‌پذیر نبود.", status=500)

    safe_page_size = max(1, min(int(page_size or 1), 100))
    safe_total = max(0, int(total))
    normalized_items: List[Dict[str, Any]] = []
    for item in items:
        code_value = str(item.get("code", "")).strip()
        part_name = str(item.get("part_name", "-")).strip() or "-"
        request_count = int(item.get("request_count", 0) or 0)
        normalized_items.append(
            {
                "code": code_value,
                "partName": part_name,
                "requestCount": request_count,
            }
        )

    total_pages = max(1, (safe_total + safe_page_size - 1) // safe_page_size)
    current_page = max(1, min(int(page or 1), total_pages))

    return {
        "items": normalized_items,
        "page": current_page,
        "pageSize": safe_page_size,
        "total": safe_total,
        "pages": total_pages,
        "range": range_key,
        "sort": sort_order,
        "search": search_value,
        "insights": _build_code_statistics_insights(range_key, search_value),
    }


def export_code_statistics_to_excel(payload: Dict[str, Any]) -> Dict[str, Any]:
    if not isinstance(payload, dict):
        raise ControlPanelError("ساختار درخواست نامعتبر است.")

    date_from = _parse_date_field(payload.get("dateFrom"), "شروع بازه")
    date_to = _parse_date_field(payload.get("dateTo"), "پایان بازه", end_of_day=True)
    if date_from and date_to and date_from > date_to:
        raise ControlPanelError("تاریخ شروع نمی‌تواند بعد از تاریخ پایان باشد.")

    raw_min_requests = payload.get("requestCount")
    min_requests: Optional[int]
    if raw_min_requests in (None, ""):
        min_requests = None
    else:
        try:
            min_requests = int(raw_min_requests)
        except Exception:
            raise ControlPanelError("حداقل تعداد درخواست باید عددی باشد.")
        if min_requests < 0:
            raise ControlPanelError("حداقل تعداد درخواست نمی‌تواند منفی باشد.")

    peak_period = str(payload.get("peakPeriod") or "day").strip().lower() or "day"
    if peak_period not in {"day", "month", "year"}:
        raise ControlPanelError("مقدار بازه پیک معتبر نیست. روز، ماه یا سال را انتخاب کنید.")

    include_mother = bool(payload.get("includeMotherCode"))
    include_product = bool(payload.get("includeProductName"))
    include_requests = bool(payload.get("includeRequestCount"))
    include_peak = bool(payload.get("includePeakPeriod"))

    file_name = _sanitize_excel_filename(payload.get("fileName"))

    try:
        records = fetch_code_statistics_for_export(
            date_from=date_from,
            date_to=date_to,
            min_request_count=min_requests,
            include_peak_period=include_peak,
            peak_period=peak_period,
        )
    except ControlPanelError:
        raise
    except Exception as exc:
        LOGGER.exception("Failed to prepare code statistics for export: %s", exc)
        raise ControlPanelError("آماده‌سازی داده‌ها برای خروجی امکان‌پذیر نبود.", status=500)

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "آمار کدها"

    peak_labels = {"day": "روز", "month": "ماه", "year": "سال"}

    headers = ["کد قطعه"]
    if include_mother:
        headers.append("کد مادر")
    if include_product:
        headers.append("نام کالا")
    if include_requests:
        headers.append("تعداد درخواست")
    if include_peak:
        headers.append(f"بازه پیک ({peak_labels.get(peak_period, 'روز')})")

    sheet.append(headers)

    rows_written = 0
    for item in records:
        normalized_code = str(item.get("code_norm") or "").strip()
        if not normalized_code:
            normalized_code = str(item.get("code_display") or "").replace("-", "").strip()

        row_values: List[Any] = [normalized_code]
        if include_mother:
            row_values.append(str(item.get("code_display") or ""))
        if include_product:
            row_values.append(str(item.get("part_name") or "-"))
        if include_requests:
            row_values.append(int(item.get("request_count") or 0))
        if include_peak:
            peak_value = str(item.get("peak_period") or "").strip()
            peak_count = int(item.get("peak_count") or 0)
            if peak_value and peak_count > 0:
                peak_text = f"{peak_value} ({peak_count} درخواست)"
            elif peak_value:
                peak_text = peak_value
            else:
                peak_text = "-"
            row_values.append(peak_text)

        sheet.append(row_values)
        rows_written += 1

    for index, header in enumerate(headers, start=1):
        column_letter = get_column_letter(index)
        max_length = len(str(header))
        for row in sheet.iter_rows(min_row=2, min_col=index, max_col=index):
            cell_value = row[0].value
            if cell_value is None:
                continue
            max_length = max(max_length, len(str(cell_value)))
        sheet.column_dimensions[column_letter].width = min(max_length + 2, 60)

    buffer = io.BytesIO()
    workbook.save(buffer)
    content = buffer.getvalue()
    encoded = base64.b64encode(content).decode("ascii")

    details = {
        "rows": rows_written,
        "minRequests": min_requests,
        "includePeak": include_peak,
        "peakPeriod": peak_period,
    }
    if date_from:
        details["dateFrom"] = date_from.isoformat()
    if date_to:
        details["dateTo"] = date_to.isoformat()

    try:
        _append_audit_event("دریافت خروجی آمار کدها", details=details)
    except Exception:
        LOGGER.debug("Failed to record export audit event", exc_info=True)

    return {
        "fileName": file_name,
        "mimeType": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        "content": encoded,
        "size": len(content),
    }


def refresh_code_stat_names(*, limit: Optional[int] = None) -> Dict[str, Any]:
    try:
        safe_limit = int(limit) if limit is not None else None
    except Exception:
        safe_limit = None

    try:
        mapping = get_inventory_name_map(refresh=True)
    except Exception as exc:
        LOGGER.warning("Failed to refresh inventory name map: %s", exc)
        raise ControlPanelError("بازخوانی نام قطعه‌ها با خطا مواجه شد.", status=500)

    updated = len(mapping)
    if updated > 0:
        _append_audit_event(
            "بازخوانی نام قطعه‌ها",
            details=f"{updated:,} نام",
        )

    limit_value = safe_limit if safe_limit is not None else updated
    return {"updated": int(updated), "limit": limit_value}


def get_commands() -> List[Dict[str, Any]]:
    commands = _load_json_setting(COMMANDS_KEY, DEFAULT_COMMANDS)
    if not isinstance(commands, list):
        commands = DEFAULT_COMMANDS
    return commands


def create_command(payload: Dict[str, Any]) -> Dict[str, Any]:
    command = (payload.get("command") or "").strip()
    if not command:
        raise ControlPanelError("دستور نمی‌تواند خالی باشد.")
    if not command.startswith("/"):
        raise ControlPanelError("دستور باید با / آغاز شود.")

    description = (payload.get("description") or "").strip()
    enabled = bool(payload.get("enabled", True))

    commands = get_commands()
    if any(item.get("command") == command for item in commands):
        raise ControlPanelError("این دستور قبلاً ثبت شده است.")

    new_item = {
        "id": f"cmd-{int(datetime.now().timestamp() * 1000)}",
        "command": command,
        "description": description,
        "enabled": enabled,
        "lastUsedISO": None,
    }
    commands.insert(0, new_item)
    _save_json_setting(COMMANDS_KEY, commands)
    _append_audit_event(
        "ثبت دستور جدید",
        details=f"{command} (شناسه {new_item['id']})",
    )
    return new_item


def update_command(command_id: str, payload: Dict[str, Any]) -> Dict[str, Any]:
    commands = get_commands()
    for idx, item in enumerate(commands):
        if item.get("id") == command_id:
            updated = {**item}
            if "command" in payload:
                value = (payload.get("command") or "").strip()
                if not value:
                    raise ControlPanelError("دستور نمی‌تواند خالی باشد.")
                if not value.startswith("/"):
                    raise ControlPanelError("دستور باید با / آغاز شود.")
                updated["command"] = value
            if "description" in payload:
                updated["description"] = (payload.get("description") or "").strip()
            if "enabled" in payload:
                updated["enabled"] = bool(payload.get("enabled"))
            commands[idx] = updated
            _save_json_setting(COMMANDS_KEY, commands)
            _append_audit_event(
                "ویرایش دستور",
                details=f"{updated['command']} (شناسه {command_id})",
            )
            return updated
    raise ControlPanelError("دستور مورد نظر یافت نشد.", status=404)


def delete_command(command_id: str) -> Dict[str, Any]:
    commands = get_commands()
    new_commands = [item for item in commands if item.get("id") != command_id]
    if len(new_commands) == len(commands):
        raise ControlPanelError("دستور مورد نظر یافت نشد.", status=404)
    _save_json_setting(COMMANDS_KEY, new_commands)
    _append_audit_event("حذف دستور", details=f"شناسه {command_id}")
    return {"success": True}


def get_blocklist() -> List[Dict[str, Any]]:
    entries: List[Dict[str, Any]] = []
    try:
        records = get_blacklist_with_meta()
        if not records:
            # Fall back to legacy behaviour
            records = [
                {"user_id": user_id, "created_at": None}
                for user_id in get_blacklist()
            ]
    except Exception as exc:
        LOGGER.warning("Failed to fetch blocklist metadata: %s", exc)
        raise ControlPanelError("دریافت لیست مسدود امکان‌پذیر نبود.", status=500)

    meta = _load_blocklist_meta()
    changed = False
    seen_keys = set()

    for record in records:
        user_id = record.get("user_id")
        if user_id is None:
            continue
        created_iso = record.get("created_at")
        key = str(user_id)
        if created_iso in ("", None):
            created_iso = meta.get(key)
        else:
            if meta.get(key) != created_iso:
                meta[key] = created_iso
                changed = True
        if created_iso in ("", None):
            created_iso = None
        entries.append(
            {
                "id": str(user_id),
                "userId": int(user_id),
                "platform": "telegram",
                "phoneOrUser": str(user_id),
                "reason": None,
                "createdAtISO": created_iso,
            }
        )
        seen_keys.add(key)

    # Clean up stale metadata for users no longer present
    stale = set(meta.keys()) - seen_keys
    if stale:
        for key in stale:
            meta.pop(key, None)
        changed = True

    if changed:
        _save_blocklist_meta(meta)

    return entries


def add_block_item(payload: Dict[str, Any]) -> Dict[str, Any]:
    raw_user_id = payload.get("userId") or payload.get("phoneOrUser")
    if raw_user_id is None:
        raise ControlPanelError("شناسه کاربر ضروری است.")
    try:
        user_id = int(str(raw_user_id).strip())
    except Exception:
        raise ControlPanelError("شناسه کاربر باید عددی باشد.")
    try:
        add_to_blacklist(user_id)
    except Exception as exc:
        LOGGER.warning("Failed to add user %s to blacklist: %s", user_id, exc)
        raise ControlPanelError("افزودن به لیست مسدود با خطا مواجه شد.", status=500)

    created_iso = None
    try:
        for record in get_blacklist_with_meta():
            if record.get("user_id") == user_id:
                created_iso = record.get("created_at")
                break
    except Exception:
        created_iso = None

    if not created_iso:
        created_iso = _now_iso()

    _remember_blocklist_timestamp(user_id, created_iso)

    _append_audit_event("افزودن به لیست مسدود", details=f"کاربر {user_id}")
    return {
        "id": str(user_id),
        "userId": user_id,
        "platform": "telegram",
        "phoneOrUser": str(user_id),
        "reason": None,
        "createdAtISO": created_iso,
    }


def remove_block_item(item_id: str) -> Dict[str, Any]:
    try:
        user_id = int(str(item_id).strip())
    except Exception:
        raise ControlPanelError("شناسه نامعتبر است.")
    try:
        remove_from_blacklist(user_id)
    except Exception as exc:
        LOGGER.warning("Failed to remove user %s from blacklist: %s", user_id, exc)
        raise ControlPanelError("حذف از لیست مسدود امکان‌پذیر نبود.", status=500)
    _forget_blocklist_timestamp(user_id)
    _append_audit_event("حذف از لیست مسدود", details=f"کاربر {user_id}")
    return {"success": True}


def get_settings() -> Dict[str, Any]:
    status = _build_status_snapshot()
    operations = status.get("operations", {})
    return {
        "timezone": status["workingHours"]["timezone"],
        "weekly": status["workingHours"]["weekly"],
        "platforms": status["platforms"],
        "lunchBreak": operations.get("lunchBreak", {"start": None, "end": None}),
        "queryLimit": operations.get("queryLimit"),
        "deliveryInfo": operations.get(
            "delivery",
            {"before": "", "after": "", "changeover": None},
        ),
        "dataSource": status.get("dataSource", "live"),
    }


def get_audit_log(*, page: int = 1, page_size: int = 20) -> Dict[str, Any]:
    try:
        page_int = int(page)
    except Exception:
        page_int = 1
    if page_int <= 0:
        page_int = 1

    try:
        size_int = int(page_size)
    except Exception:
        size_int = 50
    size_int = max(1, min(size_int, 200))

    offset = (page_int - 1) * size_int

    try:
        entries, total = fetch_audit_log_entries(limit=size_int, offset=offset)
        total_pages = max(1, (total + size_int - 1) // size_int) if total else 1
        if page_int > total_pages:
            page_int = total_pages
            offset = (page_int - 1) * size_int
            entries, total = fetch_audit_log_entries(limit=size_int, offset=offset)
        return {
            "items": entries,
            "total": total,
            "page": page_int,
            "pageSize": size_int,
            "pages": max(1, (total + size_int - 1) // size_int) if total else 1,
            "dataSource": "live",
        }
    except Exception:
        LOGGER.debug("Falling back to settings-based audit log entries", exc_info=True)

    entries = _load_audit_log_fallback()
    total = len(entries)
    total_pages = max(1, (total + size_int - 1) // size_int)
    if page_int > total_pages:
        page_int = total_pages
    start = max(0, (page_int - 1) * size_int)
    end = start + size_int
    sliced = entries[start:end]
    return {
        "items": sliced,
        "total": total,
        "page": page_int,
        "pageSize": size_int,
        "pages": total_pages,
        "dataSource": "fallback",
    }


def update_settings(payload: Dict[str, Any]) -> Dict[str, Any]:
    changes: List[str] = []
    timezone_value = payload.get("timezone")
    if timezone_value is not None:
        timezone_clean = timezone_value.strip() or "Asia/Tehran"
        set_setting(TIMEZONE_KEY, timezone_clean)
        changes.append("منطقه زمانی")

    weekly = payload.get("weekly")
    if weekly is not None:
        if not isinstance(weekly, list):
            raise ControlPanelError("ساختار ساعات کاری نامعتبر است.")

        normalized = _normalize_weekly_payload(weekly)
        try:
            current = {int(item.get("day")): item for item in fetch_working_hours_entries()}
        except Exception:
            current = {}
        for entry in normalized:
            current[int(entry["day"])] = entry
        try:
            save_working_hours_entries(current.values())
        except Exception:
            LOGGER.exception("Failed to persist working hours entries")
            raise ControlPanelError(
                "ذخیره‌سازی ساعات کاری امکان‌پذیر نبود. دوباره تلاش کنید.",
                status=500,
            )
        try:
            _sync_legacy_working_settings(current.values())
        except Exception:
            LOGGER.debug("Failed to sync legacy working hour settings", exc_info=True)
        try:
            runtime.refresh_working_hours_cache()
        except Exception:
            LOGGER.debug("Failed to refresh runtime working hours cache", exc_info=True)
        changes.append("ساعات کاری")

    platforms = payload.get("platforms")
    if isinstance(platforms, dict):
        current = _load_platform_settings(True)
        normalized = _merge_platform_flags(current, platforms)
        _save_json_setting(PLATFORMS_KEY, normalized)
        try:
            active = _is_globally_enabled()
            effective = _load_platform_settings(active)
            runtime.apply_platform_states(effective, active=active)
        except Exception:
            LOGGER.debug("Skipping runtime platform sync due to runtime error.", exc_info=True)
        changes.append("پلتفرم‌ها")

    if "lunchBreak" in payload:
        lunch_break = payload.get("lunchBreak") or {}
        if not isinstance(lunch_break, dict):
            raise ControlPanelError("ساختار بازه ناهار نامعتبر است.")
        start = _normalize_time(lunch_break.get("start"), "شروع ناهار")
        end = _normalize_time(lunch_break.get("end"), "پایان ناهار")
        set_setting(LUNCH_START_KEY, start)
        set_setting(LUNCH_END_KEY, end)
        changes.append("استراحت ناهار")

    if "queryLimit" in payload:
        limit_value = payload.get("queryLimit")
        if limit_value in (None, ""):
            set_setting(QUERY_LIMIT_KEY, "")
        else:
            try:
                limit_int = int(str(limit_value).strip())
            except Exception:
                raise ControlPanelError("محدودیت استعلام باید عددی باشد.")
            if limit_int < 0:
                raise ControlPanelError("محدودیت استعلام نمی‌تواند منفی باشد.")
            set_setting(QUERY_LIMIT_KEY, str(limit_int))
        changes.append("محدودیت استعلام")

    if "deliveryInfo" in payload:
        delivery_info = payload.get("deliveryInfo") or {}
        if not isinstance(delivery_info, dict):
            raise ControlPanelError("ساختار تنظیمات تحویل نامعتبر است.")
        before_text = str(delivery_info.get("before") or "").strip()
        after_text = str(delivery_info.get("after") or "").strip()
        changeover_value = _normalize_time(
            delivery_info.get("changeover"), "ساعت تغییر متن"
        )
        set_setting(DELIVERY_BEFORE_KEY, before_text)
        set_setting(DELIVERY_AFTER_KEY, after_text)
        set_setting(CHANGEOVER_KEY, changeover_value)
        changes.append("اطلاعات تحویل")

    if changes:
        unique_changes = []
        for item in changes:
            if item not in unique_changes:
                unique_changes.append(item)
        details = "، ".join(unique_changes)
        _append_audit_event("به‌روزرسانی تنظیمات", details=details)

    return get_settings()


def get_private_telegram_settings() -> Dict[str, Any]:
    snapshot = _snapshot_private_settings()
    delivery = snapshot.get("delivery_info") or {}

    def _extract_range(key: str) -> Dict[str, str]:
        data = snapshot.get(key) or {}
        if isinstance(data, dict):
            return {
                "start": str(data.get("start") or ""),
                "end": str(data.get("end") or ""),
            }
        return {"start": "", "end": ""}

    def _as_int_list(values: Any) -> List[int]:
        try:
            return [int(str(item).strip()) for item in (values or []) if str(item).strip()]
        except Exception:
            return []

    return {
        "enabled": bool(snapshot.get("enabled", True)),
        "dmEnabled": bool(snapshot.get("dm_enabled", snapshot.get("dmEnabled", True))),
        "apiId": snapshot.get("api_id"),
        "apiHash": snapshot.get("api_hash", ""),
        "phoneNumber": snapshot.get("phone_number", ""),
        "dataSource": snapshot.get("data_source", "sql"),
        "excelFile": snapshot.get("excel_file", ""),
        "cacheDurationMinutes": int(snapshot.get("cache_duration_minutes") or 0),
        "mainGroupId": snapshot.get("main_group_id"),
        "newGroupId": snapshot.get("new_group_id"),
        "adminGroupIds": _as_int_list(snapshot.get("admin_group_ids")),
        "secondaryGroupIds": _as_int_list(snapshot.get("secondary_group_ids")),
        "workingHours": _extract_range("working_hours"),
        "thursdayHours": _extract_range("thursday_hours"),
        "disableFriday": bool(snapshot.get("disable_friday", False)),
        "lunchBreak": _extract_range("lunch_break"),
        "queryLimit": snapshot.get("query_limit"),
        "deliveryInfo": {
            "before15": str(delivery.get("before_15") or ""),
            "after15": str(delivery.get("after_15") or ""),
        },
        "changeoverHour": str(snapshot.get("changeover_hour") or ""),
        "blacklist": _as_int_list(snapshot.get("blacklist")),
        "dataSourceOrigin": "json",
    }


def update_private_telegram_settings(payload: Dict[str, Any]) -> Dict[str, Any]:
    if not isinstance(payload, dict):
        raise ControlPanelError("ساختار تنظیمات نامعتبر است.")

    settings_ref = _ensure_private_settings()
    changes: List[str] = []

    def _set_field(key: str, value: Any, label: str) -> None:
        settings_ref[key] = value
        changes.append(label)

    if "enabled" in payload:
        _set_field("enabled", bool(payload.get("enabled")), "وضعیت ربات خصوصی")

    if "dmEnabled" in payload:
        _set_field("dm_enabled", bool(payload.get("dmEnabled")), "پاسخ خصوصی")

    if "apiId" in payload:
        try:
            api_id = int(str(payload.get("apiId")).strip())
        except Exception:
            raise ControlPanelError("API ID باید عددی باشد.")
        if api_id <= 0:
            raise ControlPanelError("API ID باید بزرگ‌تر از صفر باشد.")
        _set_field("api_id", api_id, "API ID")

    if "apiHash" in payload:
        api_hash = str(payload.get("apiHash") or "").strip()
        if not api_hash:
            raise ControlPanelError("API Hash نمی‌تواند خالی باشد.")
        _set_field("api_hash", api_hash, "API Hash")

    if "phoneNumber" in payload:
        phone = str(payload.get("phoneNumber") or "").strip()
        if not phone:
            raise ControlPanelError("شماره تلفن نمی‌تواند خالی باشد.")
        _set_field("phone_number", phone, "شماره تماس")

    if "dataSource" in payload:
        source = str(payload.get("dataSource") or "").strip().lower() or "sql"
        if source not in {"sql", "excel"}:
            raise ControlPanelError("منبع داده نامعتبر است. فقط sql یا excel مجاز است.")
        _set_field("data_source", source, "منبع داده")

    if "excelFile" in payload:
        excel_file = str(payload.get("excelFile") or "").strip()
        _set_field("excel_file", excel_file, "فایل اکسل")

    if "cacheDurationMinutes" in payload:
        try:
            cache_minutes = int(str(payload.get("cacheDurationMinutes", 0)).strip() or 0)
        except Exception:
            raise ControlPanelError("مدت کش باید عددی باشد.")
        if cache_minutes < 0:
            raise ControlPanelError("مدت کش نمی‌تواند منفی باشد.")
        _set_field("cache_duration_minutes", cache_minutes, "مدت کش")

    if "mainGroupId" in payload:
        try:
            main_group = int(str(payload.get("mainGroupId")).strip())
        except Exception:
            raise ControlPanelError("شناسه گروه اصلی نامعتبر است.")
        _set_field("main_group_id", main_group, "گروه اصلی")

    if "newGroupId" in payload:
        try:
            new_group = int(str(payload.get("newGroupId")).strip())
        except Exception:
            raise ControlPanelError("شناسه گروه جدید نامعتبر است.")
        _set_field("new_group_id", new_group, "گروه جدید")

    if "adminGroupIds" in payload:
        admin_groups = _normalize_group_list(payload.get("adminGroupIds"), "گروه‌های مدیریت")
        _set_field("admin_group_ids", admin_groups, "گروه‌های مدیریت")

    if "secondaryGroupIds" in payload:
        secondary_groups = _normalize_group_list(
            payload.get("secondaryGroupIds"), "گروه‌های فرعی"
        )
        _set_field("secondary_group_ids", secondary_groups, "گروه‌های فرعی")

    if "workingHours" in payload:
        working = _normalize_time_range(
            payload.get("workingHours"), "شروع ساعات کاری", "پایان ساعات کاری"
        )
        _set_field("working_hours", working, "ساعات کاری")

    if "thursdayHours" in payload:
        thursday = _normalize_time_range(
            payload.get("thursdayHours"), "شروع پنج‌شنبه", "پایان پنج‌شنبه"
        )
        _set_field("thursday_hours", thursday, "ساعات پنج‌شنبه")

    if "disableFriday" in payload:
        _set_field("disable_friday", bool(payload.get("disableFriday")), "وضعیت جمعه")

    if "lunchBreak" in payload:
        lunch = _normalize_time_range(
            payload.get("lunchBreak"), "شروع ناهار", "پایان ناهار"
        )
        _set_field("lunch_break", lunch, "استراحت ناهار")

    if "queryLimit" in payload:
        limit_value = payload.get("queryLimit")
        if limit_value in (None, ""):
            if "query_limit" in settings_ref:
                del settings_ref["query_limit"]
                changes.append("محدودیت استعلام")
        else:
            try:
                limit_int = int(str(limit_value).strip())
            except Exception:
                raise ControlPanelError("محدودیت استعلام باید عددی باشد.")
            if limit_int < 0:
                raise ControlPanelError("محدودیت استعلام نمی‌تواند منفی باشد.")
            _set_field("query_limit", limit_int, "محدودیت استعلام")

    if "deliveryInfo" in payload:
        delivery_payload = payload.get("deliveryInfo") or {}
        if not isinstance(delivery_payload, dict):
            raise ControlPanelError("ساختار متن تحویل نامعتبر است.")
        delivery = settings_ref.setdefault("delivery_info", {})
        delivery["before_15"] = str(delivery_payload.get("before15") or "").strip()
        delivery["after_15"] = str(delivery_payload.get("after15") or "").strip()
        changes.append("پیام تحویل")

    if "changeoverHour" in payload:
        changeover = _normalize_time(payload.get("changeoverHour"), "ساعت تغییر متن")
        _set_field("changeover_hour", changeover, "ساعت تغییر متن")

    if "blacklist" in payload:
        blacklist = _normalize_group_list(payload.get("blacklist"), "لیست سیاه")
        _set_field("blacklist", blacklist, "لیست سیاه خصوصی")

    try:
        _private_save_settings()
    except Exception as exc:  # pragma: no cover - file I/O safety
        LOGGER.error("Failed to persist private Telegram settings: %s", exc)
        raise ControlPanelError("ذخیره تنظیمات تلگرام خصوصی امکان‌پذیر نبود.", status=500)

    if changes:
        summary = "، ".join(sorted(set(changes)))
        _append_audit_event(
            "به‌روزرسانی تنظیمات تلگرام خصوصی",
            details=summary,
        )

    return get_private_telegram_settings()


def toggle_bot(active: bool) -> Dict[str, Any]:
    set_setting("enabled", "true" if active else "false")
    try:
        platforms = _load_platform_settings(active)
        runtime.apply_platform_states(platforms, active=active)
    except Exception:
        LOGGER.debug("Skipping runtime platform sync due to runtime error.", exc_info=True)
    status = _build_status_snapshot()
    _append_audit_event(
        "تغییر وضعیت ربات",
        details="فعال" if active else "غیرفعال",
    )
    return status


def get_platform_snapshot() -> Tuple[bool, Dict[str, bool]]:
    enabled = _is_globally_enabled()
    platforms = _load_platform_settings(enabled)
    return enabled, platforms


def invalidate_cache() -> Dict[str, Any]:
    try:
        # refresh_inventory_cache_once is async
        import asyncio

        asyncio.run(refresh_inventory_cache_once())
    except Exception as exc:
        LOGGER.warning("Failed to refresh inventory cache: %s", exc)
        raise ControlPanelError("به‌روزرسانی کش با خطا مواجه شد.", status=500)

    if _private_refresh_cache:
        try:
            updated = _private_refresh_cache()
            if not updated:
                LOGGER.info("Private Telegram cache refresh returned no data.")
        except Exception as exc:
            LOGGER.warning("Failed to refresh private Telegram cache: %s", exc)
            raise ControlPanelError(
                "به‌روزرسانی کش تلگرام خصوصی با خطا مواجه شد.", status=500
            )

    timestamp = _now_iso()
    set_setting(CACHE_KEY, timestamp)
    _append_audit_event("به‌روزرسانی کش کالا", details=timestamp)
    return {"lastUpdatedISO": timestamp}


def get_health() -> Dict[str, Any]:
    return {
        "status": "ok",
        "time": _now_iso(),
    }
